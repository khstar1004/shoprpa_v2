import os
from collections.abc import Generator
from typing import Any

from astronverse.scheduler.logger import logger
from openpyxl import Workbook, load_workbook


class ExcelService:
    """Excel 파일서비스"""

    def __init__(self, resource_dir: str):
        """
         Excel 서비스

        Args:
            resource_dir: 디렉터리
        """
        self.resource_dir = resource_dir

    def get_file_path(self, filename: str) -> str:
        """
        가져오기 Excel 파일의경로

        Args:
            filename: 파일이름(아니오이름)

        Returns:
            의파일 경로
        """
        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        return os.path.join(self.resource_dir, filename)

    def file_exists(self, filename: str) -> bool:
        """
        조회파일여부존재함

        Args:
            filename: 파일이름

        Returns:
            파일여부존재함
        """
        file_path = self.get_file_path(filename)
        return os.path.exists(file_path)

    def create_file(self, filename: str) -> str:
        """
        생성빈 Excel 파일

        Args:
            filename: 파일이름

        Returns:
            생성의파일 경로
        """
        file_path = self.get_file_path(filename)

        # 확인디렉터리존재함
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        # 생성빈
        wb = Workbook()
        wb.save(file_path)
        wb.close()

        logger.info(f"Created Excel file: {file_path}")
        return file_path

    def read_file_stream(self, filename: str) -> Generator[dict]:
        """
        방식가져오기 Excel 파일, 행반환데이터

        Args:
            filename: 파일이름

        Yields:
            매행데이터의딕셔너리, 형식로 {"sheet": str, "row": int, "data": list}
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path, read_only=True, data_only=True)

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 전송 sheet 열기 파일
                yield {
                    "type": "sheet_start",
                    "sheet": sheet_name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                }

                # 행가져오기데이터
                row_num = 0
                for row in ws.iter_rows(values_only=True):
                    row_num += 1
                    # 를셀값변환로가능순서열의형식
                    row_data = [self._serialize_cell_value(cell) for cell in row]
                    yield {
                        "type": "row",
                        "sheet": sheet_name,
                        "row": row_num,
                        "data": row_data,
                    }

                # 전송 sheet 결과파일
                yield {
                    "type": "sheet_end",
                    "sheet": sheet_name,
                }

        finally:
            wb.close()

        # 전송완료파일
        yield {
            "type": "complete",
            "filename": filename,
        }

    def read_file(self, filename: str) -> dict:
        """
        일가져오기 개 Excel 파일

        Args:
            filename: 파일이름

        Returns:
            패키지모든데이터의딕셔너리
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path, read_only=True, data_only=False)

        try:
            result = {
                "filename": filename,
                "sheets": [],
                "active_sheet": wb.active.title if wb.active else None,
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = {
                    "name": sheet_name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                    "data": [],
                }

                for row in ws.iter_rows(values_only=True):
                    row_data = [self._serialize_cell_value(cell) for cell in row]
                    sheet_data["data"].append(row_data)

                result["sheets"].append(sheet_data)

            return result

        finally:
            wb.close()

    def write_file(self, filename: str, data: dict) -> None:
        """
        입력데이터까지 Excel 파일

        Args:
            filename: 파일이름
            data: 입력할의데이터, 형식로 {"sheets": [{"name": str, "data": list[list]}]}
        """
        file_path = self.get_file_path(filename)

        # 확인디렉터리존재함
        dir_path = os.path.dirname(file_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path, exist_ok=True)

        wb = Workbook()

        # 삭제생성의 sheet
        if wb.active:
            wb.remove(wb.active)

        sheets = data.get("sheets", [])
        if not sheets:
            # 결과가있음데이터, 적음생성일개빈 sheet
            wb.create_sheet("Sheet1")
        else:
            for sheet_info in sheets:
                sheet_name = sheet_info.get("name", "Sheet1")
                ws = wb.create_sheet(sheet_name)

                sheet_data = sheet_info.get("data", [])
                for row_idx, row_data in enumerate(sheet_data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)

        #  sheet
        active_sheet = data.get("active_sheet")
        if active_sheet and active_sheet in wb.sheetnames:
            wb.active = wb[active_sheet]
        elif wb.sheetnames:
            wb.active = wb[wb.sheetnames[0]]

        wb.save(file_path)
        wb.close()

        logger.info(f"Saved Excel file: {file_path}")

    def update_cells(self, filename: str, updates: list[dict]) -> None:
        """
        업데이트지정셀의값

        Args:
            filename: 파일이름
            updates: 업데이트목록, 매형식로 {"sheet": str, "row": int, "col": int, "value": any}
        """
        file_path = self.get_file_path(filename)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = load_workbook(file_path)

        try:
            for update in updates:
                sheet_name = update.get("sheet")
                row = update.get("row") + 1
                col = update.get("col") + 1
                value = update.get("value")

                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]

                ws.cell(row=row, column=col, value=value)

            wb.save(file_path)
            logger.info(f"Updated {len(updates)} cells in: {file_path}")

        finally:
            wb.close()

    def delete_file(self, filename: str) -> bool:
        """
        삭제 Excel 파일

        Args:
            filename: 파일이름

        Returns:
            삭제 여부성공
        """
        file_path = self.get_file_path(filename)

        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Deleted Excel file: {file_path}")
            return True

        return False

    @staticmethod
    def _serialize_cell_value(value: Any) -> Any:
        """
        를셀값변환로가능 JSON 순서열의형식

        Args:
            value: 셀기존값

        Returns:
            순서열후의값
        """
        if value is None:
            return None
        elif isinstance(value, (str, int, float, bool)):
            return value
        else:
            # 유형변환로문자열
            return str(value)