import os
import uuid
from pathlib import Path
from unittest import TestCase

import openpyxl
from astronverse.datatable.openpyxl import OpenpyxlWrapper


class TestOpenpyxl(TestCase):
    def setUp(self):
        temp_root = Path(__file__).resolve().parents[4] / "build" / "test-openpyxl"
        os.makedirs(temp_root, exist_ok=True)
        self.test_excel_path = os.path.join(str(temp_root), f"test-{uuid.uuid4().hex}.xlsx")
        self.addCleanup(self._cleanup_test_file)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["A1", "B1", "C1"])
        sheet.append([10, 20, 30])
        sheet.append([100, 200, 300])
        sheet.cell(row=10, column=1, value=5)
        sheet.cell(row=11, column=1, value=7)
        workbook.save(self.test_excel_path)
        workbook.close()

    def _cleanup_test_file(self):
        if os.path.exists(self.test_excel_path):
            os.remove(self.test_excel_path)

    def test_openpyxl_read_cell(self):
        pyxl = openpyxl.load_workbook(self.test_excel_path)
        try:
            sheet = pyxl.active
            self.assertEqual(7, sheet.cell(row=11, column=1).value)
        finally:
            pyxl.close()

    def test_openpyxl_write_cell(self):
        pyxl = openpyxl.load_workbook(self.test_excel_path)
        try:
            sheet = pyxl.active
            sheet.cell(row=1, column=1, value="Test Value")
            pyxl.save(self.test_excel_path)
        finally:
            pyxl.close()

        saved = openpyxl.load_workbook(self.test_excel_path)
        try:
            self.assertEqual("Test Value", saved.active.cell(row=1, column=1).value)
        finally:
            saved.close()

    def test_read_cell(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        self.assertEqual("A1", wrapper.read_cell(row=1, col=1))

    def test_read_row(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        self.assertEqual(["A1", "B1", "C1"], wrapper.read_row(row_index=1))

    def test_read_column(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        values = wrapper.read_column(col_index=1)
        self.assertEqual("A1", values[0])
        self.assertEqual(5, values[9])
        self.assertEqual(7, values[10])

    def test_read_area(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        self.assertEqual(
            [["A1", "B1", "C1"], [10, 20, 30], [100, 200, 300]],
            wrapper.read_range("A1:C3"),
        )

    def test_read_all(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        values = wrapper.read_effective_area()
        self.assertEqual(["A1", "B1", "C1"], values[0])
        self.assertEqual(7, values[10][0])

    def test_get_max_row(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        self.assertEqual(11, wrapper.get_max_row())

    def test_get_max_column(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        self.assertEqual(3, wrapper.get_max_column())

    def test_write_cell(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        wrapper.write_cell(row=1, col=1, value="11")
        wrapper.save()

        saved = openpyxl.load_workbook(self.test_excel_path)
        try:
            self.assertEqual("11", saved.active.cell(row=1, column=1).value)
        finally:
            saved.close()

    def test_insert_cell(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        wrapper.insert_cells(row=2, col=2, amount=1)
        self.assertIsNone(wrapper.read_cell(row=2, col=2))

    def test_write_cell_formula(self):
        wrapper = OpenpyxlWrapper(file_path=self.test_excel_path)
        wrapper.write_cell(row=12, col=1, value="=SUM(A10:A11)")
        wrapper.save()

        saved = openpyxl.load_workbook(self.test_excel_path, data_only=False)
        try:
            self.assertEqual("=SUM(A10:A11)", saved.active.cell(row=12, column=1).value)
        finally:
            saved.close()
